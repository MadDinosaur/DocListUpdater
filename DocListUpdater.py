## Imports ##
import os
import time
import openpyxl
from bs4 import BeautifulSoup

## Global variables ##
TIME_INTERVAL = 86400  # REPRESENTS HOW OLD THE FILE MODIFICATION DATE SHOULD BE FOR THE FILE TO BE ADDED TO THE LIST AS A "NEW FILE" (86400 = 24H)

FILE = "listadocs.xlsx"

WORKBOOK = openpyxl.load_workbook(FILE)

SHEET = WORKBOOK[WORKBOOK.sheetnames[0]]

HEADERS = []
for cell in tuple(SHEET.rows)[0]:
    if cell.value is None:
        NUM_MAIN_HEADERS = cell.column - 1
    HEADERS.append(cell.value)

PATHS = []
for col in SHEET.columns:
    if col[0].value == HEADERS[-2]:
        for index in range(1, len(col)):
            if col[index].value is None:
                break
            PATHS.append(col[index].value)
        break

FORBIDDEN_PATHS = []
for col in SHEET.columns:
    if col[0].value == HEADERS[-1]:
        for index in range(1, len(col)):
            if col[index].value is None:
                break
            FORBIDDEN_PATHS.append(col[index].value)
        break

TYPES_OF_DOCUMENTS = []
for col in SHEET.columns:
    if col[0].value == "Tipos":
        for index in range(1, len(col)):
            if col[index].value is None:
                break
            TYPES_OF_DOCUMENTS.append(col[index].value)
        break

FILE_EXTENSION_BLACK_LIST = ['ini', 'db', 'DS_Store']

##Methods##
def find_files():
    file_list = []

    for path in PATHS:
        for root, dir, files in os.walk(path):

            include_files = True
            for not_path in FORBIDDEN_PATHS:
                if root.find(not_path) != -1:
                    include_files = False

            if include_files:
                for filename in files:
                    file_extension = filename.split('.')[1]
                    if file_extension in FILE_EXTENSION_BLACK_LIST:
                        continue
                    else:
                        for type in TYPES_OF_DOCUMENTS:
                            if root.find(type[:-3]) != -1:
                                file_list_type = type
                                break
                        file_list_filename = filename.split(".")[0]
                        file_list_location = root + "\\" + filename

                        file_list.append([file_list_type, file_list_filename, file_list_location])

    return file_list


def find_files_by_date():
    file_list = []

    for path in PATHS:
        for root, dir, files in os.walk(path):

            include_files = True
            for not_path in FORBIDDEN_PATHS:
                if root.find(not_path) != -1:
                    include_files = False

            if include_files:
                for filename in files:

                    file_date = os.path.getmtime(root + "\\" + filename)
                    file_extension = filename.split('.')[1]
                    if file_extension in FILE_EXTENSION_BLACK_LIST:
                        continue
                    elif file_date >= time.time() - TIME_INTERVAL:
                        for type in TYPES_OF_DOCUMENTS:
                            if root.find(type[:-3]) != -1:
                                file_list_type = type
                                break
                        file_list_filename = filename.split(".")[0]
                        file_list_location = root + "\\" + filename

                        file_list.append([file_list_type, file_list_filename, file_list_location])

    return removeDuplicates(file_list)


# def write_files():
#     file_list_index = 0
#     file_list_size = len(FILE_LIST)
#     row_number = 2
#     col_number_type = HEADERS.index("Tipo") + 1
#     col_number_location = HEADERS.index("Localização") + 1
#     for row in SHEET.iter_rows(min_row=2, max_row=file_list_size + row_number, max_col=NUM_MAIN_HEADERS,
#                                values_only=True):
#         if file_list_index >= file_list_size:
#             break
#         else:
#             row_is_empty = row.count(None) == len(row)
#             if row_is_empty:
#                 SHEET.cell(row=row_number, column=col_number_type).value = FILE_LIST[file_list_index][0]
#                 SHEET.cell(row=row_number, column=col_number_location).hyperlink = FILE_LIST[file_list_index][2]
#                 SHEET.cell(row=row_number, column=col_number_location).value = FILE_LIST[file_list_index][1]
#                 file_list_index += 1
#             row_number += 1
#
#     WORKBOOK.save(FILE)


def removeDuplicates(newfiles_list):
    files_list = []
    return list(set([i for i in files_list]))


def convertToHtml():
    import HTMLConverter
    for line in FILE_LIST:
        type = line[0]
        filename = line[1]
        location = line[2]
        path = "file:" + location.replace("\\", "/")

        HTMLConverter.add_line(type, filename, location, path)
    HTMLConverter.writeToHTML()

## Method calls ##
FILE_LIST = find_files_by_date()
convertToHtml()
